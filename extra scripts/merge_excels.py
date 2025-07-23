import pandas as pd
import xlsxwriter
from pathlib import Path
import openpyxl
import os
import warnings
"""
Merges all th excel files for every subfolder of Results

Need to run at the same directory as the src folder

Refactoring-recognition-for-python
├── Results/
├── Documentation/
├── src/
├── merge_excels.py    <------
├── requirements.txt 
└── README.md
"""
def find_excel_files_in_subfolders(folder):
    excel_files = []
    releases_comp =[]
    releases_count = 0
    comparable_files_count=0
    subfolders = [f.path for f in os.scandir(folder) if f.is_dir()]
    
    for subfolder in subfolders:
        excel_found_flag =False       
        for f in os.scandir(subfolder):
            f = Path(f)
            if f.is_file():
                if f.name.endswith('.xlsx') or f.name.endswith('.xls'):
                    excel_files.append(f)
                    excel_found_flag = True
                    break
        if excel_found_flag:
            for f in os.scandir(subfolder):
                if f.name == 'Execution.txt':
                        execution_file = f
                        num_of_releases_per_project = None
                        num_of_comparable_files_per_project = None
                        releases_string = None
                        with open(execution_file, 'r') as file:
                            found_releases_flag =False
                            for line in file:
                                if line.startswith('Number of releases:'):
                                    num_of_releases_per_project =int(line.split(':')[-1])
                                    releases_count += int(line.split(':')[-1])
                                if line.startswith('Total comparable files :'):
                                    num_of_comparable_files_per_project = int(line.split(':')[-1])
                                    comparable_files_count+=int(line.split(':')[-1])
                                if found_releases_flag:
                                    releases_string =line
                                    found_releases_flag = False
                                if line.startswith("Releases:"):
                                    found_releases_flag = True
                        if num_of_releases_per_project != None and num_of_comparable_files_per_project != None and releases_string != None:   
                            releases_comp.append([num_of_releases_per_project,num_of_comparable_files_per_project,releases_string])
                            break
                                    
    print(f"Number of releases : {releases_count}")
    print(f"Number of comparable files : {comparable_files_count}")
    print(f"Number of projects : {len(excel_files)}")
    return excel_files,releases_count,comparable_files_count,releases_comp

def create_chart(categories,values,cell_location,project_name,style,workbook,worksheet_):
    chart = workbook.add_chart({"type": style})
    chart.set_style(2)    
    chart.add_series({
                    "name": f"{project_name}",
                    "categories": categories,
                    #"categories": f"=Projects!$B${start+2}:$B${end}",
                    #'values':     f"=Projects!$C${start+2}:$C${end}",
                    'values':      values,
                    'data_labels': {'value': True},
                })
    chart.set_size({'width': 576, 'height': 579.84})
    chart.set_legend({'none': True})
    chart.set_x_axis({'min': 0, 'max': 100, 'num_format': '0\%'})
    worksheet_.insert_chart(f'{cell_location}', chart)
 

def create_statistics_per_project(df,number_of_releases,number_of_comparable_files,releases_string):
    total_count = len(df)
    simplifying_method_calls= [
        'Add/Remove Parameter',
        'Rename Variable',
        'Preserve Whole Object',
        'Introduce Parameter Object',
        'Remove Setting Method',
        'Hide Method',
        'Replace ErrorCode With Exception',
        'Replace Exception With Test'
                               ]
    moving_features_between_objects = [
        'Move Method',
        'Extract Class',
        'Inline Class',
        'Move Field'
    ]
    organizing_data = [
        'Replace Array With Object',
        'Replace Magic Number With Symbolic Constant',
        'Encapsulated Collection',
        'Encapsulated Field'
    ]
    dealing_with_generalizations = [
        'Pull Up Field',
        'Push Down Field',
        'Pull Up Method',
        'Push Down Method',
        'Pull Up Constructor Body',
        'Extract Subclass',
        'Extract Superclass',
        'Collapse Hierarchy'
    ]
    categories = [
        'Simplifying Method Calls',
        'Moving Features between Objects',
        'Organizing Data',
        'Dealing with Generalizations'
    ]
    
  
    
    def release_with_the_most_refactors():
        def count_refactor_per_release(release):
            return (df['version 1'] == release).sum()
        releases_name_v1 = list(df['version 1'].unique())
        releases_name_v2 = [[release, df.loc[df['version 1'] == release, 'version 2'].iloc[0]] for release in releases_name_v1]
        releases_name = list(zip(releases_name_v1, releases_name_v2))
        releases_count =[ count_refactor_per_release(release[0]) for release in releases_name]
        releases = list(zip(releases_name, releases_count))
        max_release = max(releases, key=lambda x: x[1])
        return "["+max_release[0][1][0] +"] and ["+ max_release[0][1][1]+"]"
    
    
    def table_per_category(category):
        table1= []
        for refactoring in category:
            ref_count = (df['Refactor name'] == refactoring).sum()
            value = (ref_count / total_count) * 100
            table1.append(["",refactoring,value,""])
        return table1
   
    
    releases_with_most_refactors = str(release_with_the_most_refactors())
    comparable_files_count = number_of_comparable_files
    unique_versions_count =number_of_releases
    
    releases_list = releases_string
    try:
        project_name = df.loc[1, "Project name"]   
        if ","  in project_name:
            project_name = project_name.split(",")[1].replace('"',"").replace(")","").strip()
        if "_"  in project_name:
            github_link = "https://github.com/"+project_name.split("_")[0]+"/"+project_name.split("_")[1]
        else: github_link = "-"
    except  KeyError:
        project_name = "if you see this something went wrong"
        github_link = "-"
    
    table =[]
    value = ((df['Category'] == 'Simplifying Method Calls').sum() / total_count) * 100
    table.append(["Simplifying Method Calls","","",value])
    table.extend(table_per_category(simplifying_method_calls))
    value = ((df['Category'] == 'Moving Features between Objects').sum() / total_count) * 100
    table.append(["Moving Features between Objects","","",value])
    table.extend(table_per_category(moving_features_between_objects))
    value = ((df['Category'] == 'Organizing Data').sum() / total_count) * 100
    table.append(["Organizing Data","","",value])
    table.extend(table_per_category(organizing_data))
    value= ((df['Category'] == 'Dealing With Generalizations').sum() / total_count) * 100
    table.append(["Dealing With Generalizations","","",value])
    table.extend(table_per_category(dealing_with_generalizations))
    table1 =[]
    table1.append(["Project name",project_name])
    table1.append(["Github link",github_link])
    table1.append(["Total releases",unique_versions_count])
    table1.append(["Releases",releases_list])
    table1.append(["Total comparable files",comparable_files_count])
    table1.append(["Total refactors",total_count])
    table1.append(["Releases with the most refactors",releases_with_most_refactors])
    
            
    return table,table1,project_name
                  
def merge_excel_files(folder_path, output_file):
    # List all Excel files in the folder
    excel_files,releases_count,comparable_files_count,releases_and_comparable_file_per_project_list = find_excel_files_in_subfolders(folder_path)
    if not excel_files:
        print("No Excel files found in the specified folder.")
        return
   
    dfs = []
    stats= []
    for file,releases_and_comparable_file in zip(excel_files,releases_and_comparable_file_per_project_list):
        number_of_releases = releases_and_comparable_file[0]
        number_of_comparable_files = releases_and_comparable_file[1]
        releases_string = releases_and_comparable_file[2]
        file_path = Path(folder_path) / file
        workbook_read = openpyxl.load_workbook(file_path)
        worksheet_read = workbook_read.active  
        data = []
        for row in worksheet_read.iter_rows(values_only=True):
            data.append(row)
        df = pd.DataFrame(data[1:], columns=data[0]) 
        table,table1,project_name = create_statistics_per_project(df,number_of_releases,number_of_comparable_files,releases_string)
        stats.append([table,table1,project_name])
        dfs.append(df)
        
        
    merged_data = pd.concat(dfs, ignore_index=True)
    
    sort_column = "Refactor name"
    if sort_column:
        merged_data.sort_values(by=sort_column, inplace=True)

    if 'Link to html comparison file' in merged_data.columns:
        merged_data.drop(columns=['Link to html comparison file'], inplace=True)
 
    # Create a new Excel workbook
    workbook = xlsxwriter.Workbook(output_file)
    centered_format = workbook.add_format({'align': 'center'})
    left_format = workbook.add_format({'align': 'left'})
    number_percent_format = workbook.add_format({'num_format': '0.000\%', 'bold': True, 'align': 'center'})
    bold = workbook.add_format({'align': 'center', 'bold': True})
    
    worksheet = workbook.add_worksheet("Refactors")

    # Write the data
    for row_num, row_data in enumerate(merged_data.values):
        for col_num, cell_data in enumerate(row_data):
            worksheet.write(row_num, col_num, cell_data)

    worksheet.set_column('A:A',45)
    worksheet.set_column('B:B',32)
    worksheet.set_column('C:C',20, centered_format)
    worksheet.set_column('D:E',10, centered_format)
    worksheet.set_column('F:G',45, centered_format)
    worksheet.set_column('I:K',12)

    
    # Define a table
    end_row = len(merged_data)
    table_range = f'A1:K{end_row}'
    worksheet.add_table(table_range, {'autofilter': False,
                            'style': 'Table Style Medium 9',
                            'header_row': True,
                                'columns': [{'header': 'Refactor name'},
                                            {'header': 'Category'},
                                            {'header': 'Project name'},
                                            {'header': 'version 1'},
                                            {'header': 'version 2'},
                                            {'header': 'File 1 Path'},
                                            {'header': 'File 2 Path'},
                                            {'header': 'File 1 start line'},
                                            {'header': 'File 1 end line'},
                                            {'header': 'File 2 start line'},
                                            {'header': 'File 2 end line'},
                                            ]})

    # Add a new worksheet
    worksheet2 = workbook.add_worksheet('General stats') 
    table1 =[]
    table1.append(["Simplifying Method Calls","","",f"=COUNTIF('Refactors'!B2:B{end_row},A2)","=D2/B36 * 100"])
    table1.append(["","Add/Remove Parameter",f"=COUNTIF('Refactors'!A2:A{end_row},B3)","","=C3/B36 * 100"])
    table1.append(["","Rename Variable",f"=COUNTIF('Refactors'!A2:A{end_row},B4)","","=C4/B36 * 100"])
    table1.append(["","Preserve Whole Object",f"=COUNTIF('Refactors'!A2:A{end_row},B5)","","=C5/B36 * 100"])
    table1.append(["","Introduce Parameter Object",f"=COUNTIF('Refactors'!A2:A{end_row},B6)","","=C6/B36 * 100"])
    table1.append(["","Remove Setting Method",f"=COUNTIF('Refactors'!A2:A{end_row},B7)","","=C7/B36 * 100"])
    table1.append(["","Hide Method",f"=COUNTIF('Refactors'!A2:A{end_row},B8)","","=C8/B36 * 100"])
    table1.append(["","Replace ErrorCode With Exception",f"=COUNTIF('Refactors'!A2:A{end_row},B9)","","=C9/B36 * 100"])
    table1.append(["","Replace Exception With Test",f"=COUNTIF('Refactors'!A2:A{end_row},B10)","","=C10/B36 * 100"])
    
    table1.append(["Moving Features between Objects","","",f"=COUNTIF('Refactors'!B2:B{end_row},A11)","=D11/B36 * 100"])
    table1.append(["","Move Method",f"=COUNTIF('Refactors'!A2:A{end_row},B12)","","=C12/B36 * 100"])
    table1.append(["","Extract Class",f"=COUNTIF('Refactors'!A2:A{end_row},B13)","","=C13/B36 * 100"])
    table1.append(["","Inline Class",f"=COUNTIF('Refactors'!A2:A{end_row},B14)","","=C14/B36 * 100"])
    table1.append(["","Move Field",f"=COUNTIF('Refactors'!A2:A{end_row},B15)","","=C15/B36 * 100"])
    
    table1.append(["Organizing Data","","",f"=COUNTIF('Refactors'!B2:B{end_row},A16)","=D16/B36 * 100"])
    table1.append(["","Replace Array With Object",f"=COUNTIF('Refactors'!A2:A{end_row},B17)","","=C17/B36 * 100"])
    table1.append(["","Replace Magic Number With Symbolic Constant",f"=COUNTIF('Refactors'!A2:A{end_row},B18)","","=C18/B36 * 100"])
    table1.append(["","Encapsulated Collection",f"=COUNTIF('Refactors'!A2:A{end_row},B19)","","=C19/B36 * 100"])
    table1.append(["","Encapsulated Field",f"=COUNTIF('Refactors'!A2:A{end_row},B20)","","=C20/B36 * 100"])
    
    table1.append(["Dealing with Generalizations","","",f"=COUNTIF('Refactors'!B2:B{end_row},A21)","=D21/B36 * 100"])
    table1.append(["","Pull Up Field",f"=COUNTIF('Refactors'!A2:A{end_row},B22)","","=C22/B36 * 100"])
    table1.append(["","Push Down Field",f"=COUNTIF('Refactors'!A2:A{end_row},B23)","","=C23/B36 * 100"])
    table1.append(["","Pull Up Method",f"=COUNTIF('Refactors'!A2:A{end_row},B24)","","=C24/B36 * 100"])
    table1.append(["","Push Down Method",f"=COUNTIF('Refactors'!A2:A{end_row},B25)","","=C25/B36 * 100"])
    table1.append(["","Pull Up Constructor Body",f"=COUNTIF('Refactors'!A2:A{end_row},B26)","","=C26/B36 * 100"])
    table1.append(["","Extract Subclass",f"=COUNTIF('Refactors'!A2:A{end_row},B27)","","=C27/B36 * 100"])
    table1.append(["","Extract Superclass",f"=COUNTIF('Refactors'!A2:A{end_row},B28)","","=C28/B36 * 100"])
    table1.append(["","Collapse Hierarchy",f"=COUNTIF('Refactors'!A2:A{end_row},B29)","","=C29/B36 * 100"])

    worksheet2.set_column('A:A',32)
    worksheet2.set_column('B:B',45)
    worksheet2.set_column('C:D',45, centered_format)
    worksheet2.set_column('E:E',20, number_percent_format)
    worksheet2.add_table(f'A1:E{29}', {'data': table1,
                            'autofilter': False,
                            'style': 'Table Style Medium 9',
                            'header_row': True,
                                'columns': [{'header': 'Category'},
                                            {'header': 'Refactor'},
                                            {'header': 'count' ,'format': bold},
                                            {'header': 'count per category' ,'format': bold},
                                            {'header': '% of total refactors'},
                                            ]})
    
    table2 =[]
    table2.append(["Projects",len(excel_files)])
    table2.append(["Releases",releases_count])
    table2.append(["Comparable files",comparable_files_count])
    table2.append(["Refactors",f"=SUM(C3:C10,C12:C15,C17:C20,C22:C29)"])
    worksheet2.add_table(f'A32:B{36}', {'data': table2,
                            'autofilter': False,
                            'style': 'Table Style Medium 9',
                            'header_row': True,
                                'columns': [{'header': 'Number of'},
                                            {'header': 'Total', 'format': bold},
                                            ]})
    
    #value = 3
    #categories = f"='General stats'!$B${value}:$B${value+7},'General stats'!$B${value+9}:$B${value+12},'General stats'!$B${value+14}:$B${value+17},'General stats'!$B${value+19}:$B${value+26}"
    #values = f"='General stats'!$E${value}:$E${value+7},'General stats'!$E${value+9}:$E${value+12},'General stats'!$E${value+14}:$E${value+17},'General stats'!$E${value+19}:$E${value+26}"
    
    categories = f"='General stats'!$A${2},'General stats'!$A${11},'General stats'!$A${16},'General stats'!$A${21}"
    values = f"='General stats'!$E${2},'General stats'!$E${11},'General stats'!$E${16},'General stats'!$E${21}"
    create_chart(categories,values,f"B{int(39)}","Refactors per category","bar",workbook,worksheet2) #refactor pie chart categories
    
    create_chart(f"='General stats'!$B${3}:$B${10}",f"='General stats'!$E${3}:$E${10}",f"A{int(69)}","Category : Simplifying Method Calls","column",workbook,worksheet2) #refactor pie chart per category
    create_chart(f"='General stats'!$B${12}:$B${15}",f"='General stats'!$E${12}:$E${15}",f"D{int(69)}","Category : Moving Features between Objects","column",workbook,worksheet2)
    create_chart(f"='General stats'!$B${17}:$B${20}",f"='General stats'!$E${17}:$E${20}",f"A{int(99)}","Category : Organizing Data","column",workbook,worksheet2)
    create_chart(f"='General stats'!$B${22}:$B${29}",f"='General stats'!$E${22}:$E${29}",f"D{int(99)}","Category : Dealing with Generalizations","column",workbook,worksheet2)
    
    projects_with_most_refactors = []
    for stats_ in stats:
        project_data = stats_[1]
        project_data_refactors = stats_[0]
        max_times_refactor = max((rf[2] for rf in project_data_refactors if rf[2]), default=0)
        refactor_occurred_the_most = next((rf[1] for rf in project_data_refactors if rf[2] == max_times_refactor), None)
        projects_with_most_refactors.append([project_data[0][1], project_data[5][1], refactor_occurred_the_most])
    projects_with_most_refactors.sort(key=lambda x: int(x[1]), reverse=True)
    
    total_projects=len(projects_with_most_refactors)
    worksheet2.add_table(f'A130:C{130+total_projects}', {'data': projects_with_most_refactors,
                            'autofilter': False,
                            'style': 'Table Style Medium 9',
                            'header_row': True,
                                'columns': [{'header': 'Project'},
                                            {'header': 'Refactors count', 'format': bold},
                                            {'header': 'Refactor name occurred the most', 'format': bold},
                                            ]})

    worksheet_ = workbook.add_worksheet("Projects") 
    wrapped_text_format = workbook.add_format({'text_wrap': True})
    worksheet_.set_column('A:A',32)
    worksheet_.set_column('B:B',45, wrapped_text_format)
    worksheet_.set_column('C:D',30, number_percent_format)
    plus_value = 29
    end =0
    start =1
    
    for stats_ in stats:
        table = stats_[0]
        table1 = stats_[1]
        project_name = stats_[2]
        if len(project_name) > 30:
            project_name = project_name[:30]
        if project_name!= "-":
            end =start +7
            worksheet_.add_table(f'A{start}:B{end}', {'data': table1,
                                'autofilter': False,
                                'style': 'Table Style Medium 9',
                                })
            start = end+1
            end +=plus_value
            worksheet_.add_table(f'A{start}:D{end}', {'data': table,
                                'autofilter': False,
                                'style': 'Table Style Medium 9',
                                'header_row': True,
                                    'columns': [{'header': 'Category'},
                                                {'header': 'Refactor'},
                                                {'header': '% of total refactors' },
                                                {'header': '% of total refactors per category' },
                                                ]})
            
            categories = f"=Projects!$A${start+1},Projects!$A${start+10},Projects!$A${start+15},Projects!$A${start+20}"
            values = f"=Projects!$D${start+1},Projects!$D${start+10},Projects!$D${start+15},Projects!$D${start+20}"
            create_chart(categories,values,f"F{int(start)}",project_name,"bar",workbook,worksheet_) 
            create_chart(f"=Projects!$B${start+2}:$B${start+9}",f"=Projects!$C${start+2}:$C${start+9}",f"P{int(start)}",f"{project_name}\nCategory : Simplifying Method Calls","column",workbook,worksheet_) #refactor pie chart per category
            create_chart(f"=Projects!$B${start+11}:$B${start+14}",f"=Projects!$C${start+11}:$C${start+14}",f"Z{int(start)}",f"{project_name}\nCategory : Moving Features between Objects","column",workbook,worksheet_)
            create_chart(f"=Projects!$B${start+16}:$B${start+19}",f"=Projects!$C${start+16}:$C${start+19}",f"AJ{int(start)}",f"{project_name}\nCategory : Organizing Data","column",workbook,worksheet_)
            create_chart(f"=Projects!$B${start+21}:$B${start+28}",f"=Projects!$C${start+21}:$C${start+28}",f"AT{int(start)}",f"{project_name}\nCategory : Dealing with Generalizations","column",workbook,worksheet_)
            
            
            start = end+4
    

        
    # Close the workbook
    workbook.close()

    print(f"Merged excels saved to {output_file}")
if __name__ == '__main__':
    current_dir = Path(__file__).resolve().parent
    folder_path = current_dir / "Results"
    output_file = current_dir / "Refactors_merged.xlsx"
    warnings.filterwarnings('ignore', message="Unknown worksheet reference")
    merge_excel_files(folder_path, output_file)
